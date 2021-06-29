import os

import xlrd
import xlwt
import openpyxl
import urllib.request
import numpy as np
import cv2


def get_image(path):
    try:
        response = urllib.request.urlopen(path)
        image = response.read()
    except:
        return None
    return image


def get_image_np(path):
    image = get_image(path)
    if image is None:
        return None
    np_array = np.fromstring(image, np.uint8)
    image = cv2.imdecode(np_array, cv2.IMREAD_COLOR)
    return image


data = openpyxl.load_workbook('./反光衣安全帽.xlsx')
sheet_name = data.get_sheet_names()
print('sheet name:', sheet_name)
sheet1 = data.get_sheet_by_name(sheet_name[0])
rows = sheet1.max_row
cols = sheet1.max_column
print("rows:{},cols:{}".format(rows, cols))

save_dir = './download_imgs'
img_paths = []
for i in range(rows):
    img_path = sheet1.cell(i + 1, 12).value
    img_path = os.path.join("https://site.gd-ib.cn/site/file/", img_path)
    img_paths.append(img_path)

for i, item in enumerate(img_paths):
    print("img_path:", item)
    try:
        response = urllib.request.urlopen(item)
        image = response.read()
        np_array = np.fromstring(image, np.uint8)
        image = cv2.imdecode(np_array, cv2.IMREAD_COLOR)
        cv2.imwrite("./download_imgs/{}.jpg".format(str(i)), image)
    except Exception as e:
        print(e)

    # img = get_image_np(item)
